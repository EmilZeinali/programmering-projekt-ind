from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
wb = load_workbook("programmering.xlsx")
ws = wb.active
A = 20
B = 17.5
C = 15
D = 12.5
E = 10
F = 0
subjects = ["Sam", "Matte", "Engelska", "Svenska", "Idrott", "Teknik", "Programmering"]
values = {
    "A": 20,
    "B": 17.5,
    "C": 15,
    "D": 12.5,
    "E": 10,
    "F": 0
}

for i in range(30):
    sum = 0
    for subject in subjects:
        val=input("vad fick eleven i" + subject )
        sum += values()
    average = sum/len(subjects)

